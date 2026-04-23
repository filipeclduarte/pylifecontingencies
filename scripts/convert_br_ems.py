#!/usr/bin/env python3
"""
Extract BR-EMS biometric tables from the official SUSEP Excel file and write
age/qx CSV files to src/pylifecontingencies/data/.

Requires openpyxl (dev dependency):
  pip install openpyxl

Usage:
  python scripts/convert_br_ems.py
  python scripts/convert_br_ems.py --xlsx path/to/Tabuas.xlsx
  python scripts/convert_br_ems.py --out-dir /custom/output/path
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

DEFAULT_XLSX = (
    Path(__file__).parent.parent / "Tabuas BR-EMS 2010 2015 2021-010721.xlsx"
)
DEFAULT_OUT_DIR = (
    Path(__file__).parent.parent / "src" / "pylifecontingencies" / "data"
)


def _canonical_to_key(canonical: str) -> str:
    """
    Convert the canonical table name printed in the spreadsheet (row 4) to
    a Python-friendly identifier.

    Examples
    --------
    'BR-EMSsb-v.2021-m' -> 'br_emssb_2021_m'
    'BR-EMSmt-v.2010-f' -> 'br_emsmt_2010_f'
    """
    # Strip whitespace, lower-case
    s = canonical.strip().lower()
    # Remove 'v.' version prefix and all separators (-, .)
    s = re.sub(r"[-.\s]+", "_", s)
    # Remove the "v_" fragment left after stripping "v."
    s = re.sub(r"_v_", "_", s)
    # Remove any leading/trailing underscores
    s = s.strip("_")
    return s


def _extract_single_sheet(ws, sheet_name: str) -> tuple[str, list[tuple[int, float]]]:
    """
    Extract (canonical_name, [(age, qx), ...]) from a 2015/2021 sheet.

    Layout:
      Row 4, col B  — canonical table name (e.g. 'BR-EMSsb-v.2021-m')
      Row 6         — header: Idade | qx | -IC(95%) | +IC(95%) | lx | ex
      Row 7+        — data
    """
    rows = list(ws.iter_rows(values_only=True))

    # Canonical name is in row 4 (index 3), column B (index 1)
    canonical = str(rows[3][1]).strip()

    data: list[tuple[int, float]] = []
    for row in rows[6:]:  # data starts at row 7 (0-indexed row 6)
        if row[1] is None:
            break
        age = row[1]
        qx = row[2]
        if not isinstance(age, (int, float)):
            continue
        if not isinstance(qx, (int, float)):
            continue
        data.append((int(age), float(qx)))
        if float(qx) >= 1.0:
            break  # terminal age reached; drop any duplicate qx=1 rows

    return canonical, data


def _extract_2010_sheet(ws) -> dict[str, list[tuple[int, float]]]:
    """
    Extract four tables from the combined 2010 sheet.

    Layout:
      Row 4, cols C-F  — canonical names
      Row 5            — header row ('Idade', 'qx', 'qx', 'qx', 'qx')
      Row 6+           — data
    """
    rows = list(ws.iter_rows(values_only=True))

    # Row 4 (index 3): col C=index 2, D=3, E=4, F=5
    canonicals = [str(rows[3][i]).strip() for i in (2, 3, 4, 5)]

    tables: dict[str, list[tuple[int, float]]] = {c: [] for c in canonicals}

    done = set()
    for row in rows[5:]:  # data starts at row 6 (0-indexed row 5)
        age = row[1]
        if not isinstance(age, (int, float)):
            continue
        for i, canonical in enumerate(canonicals):
            if canonical in done:
                continue
            qx = row[2 + i]
            if isinstance(qx, (int, float)):
                tables[canonical].append((int(age), float(qx)))
                if float(qx) >= 1.0:
                    done.add(canonical)

    return tables


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to the BR-EMS Excel workbook",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=DEFAULT_OUT_DIR,
        help="Output directory for CSV files",
    )
    args = parser.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    if not args.xlsx.exists():
        print(f"ERROR: Excel file not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    args.out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Reading: {args.xlsx}")
    print(f"Output:  {args.out_dir}\n")

    wb = openpyxl.load_workbook(str(args.xlsx), read_only=True, data_only=True)

    tables: dict[str, list[tuple[int, float]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        stripped = sheet_name.strip()

        if stripped == "Vigência":
            continue

        if "2010" in stripped:
            extracted = _extract_2010_sheet(ws)
            tables.update(extracted)
        else:
            canonical, data = _extract_single_sheet(ws, sheet_name)
            tables[canonical] = data

    import csv

    written = []
    for canonical, data in sorted(tables.items()):
        key = _canonical_to_key(canonical)
        out_path = args.out_dir / f"{key}.csv"
        with open(out_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["age", "qx"])
            writer.writerows(data)
        print(f"  {key:30s}  ({len(data)} ages, {canonical})")
        written.append(key)

    print(f"\nDone. {len(written)} tables written to {args.out_dir}")


if __name__ == "__main__":
    main()
